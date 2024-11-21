from openpyxl import load_workbook

# lê pasta de trabalho e planilha
wb = load_workbook('data/datapivot_table.xlsx')
sheet = wb['Relatorio']
# print(sheet)

# acessando um valor especifico
print(sheet['B3'].value)

# interando sobre as linhas
# for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=3):
#     # get the year and value from the spreadsheet
#     year = row[0].value
#     value = row[1].value
#     # print the values
#     print(year, value)

for i in range(2,6):
    ano = sheet["A%s" % i].value
    frabricante = sheet["B%s" % i].value
    valor = sheet["C%s" % i].value
    print("{0} o Aston Martin vendeu {1} e o Bentley vendeu {2}".format(ano, frabricante, valor))